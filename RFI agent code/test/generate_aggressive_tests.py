"""
Aggressive tests designed to BREAK the RFI pipeline.
These target specific weaknesses identified in code review.
"""

import os
import struct
from openpyxl import Workbook

OUT_DIR = os.path.dirname(os.path.abspath(__file__))


def test1_numeric_only_sheets_bug():
    """
    TEST 1 (Easy - known bug): Sheets with ONLY numeric data should NOT
    be parsed as containing questions. The parser's _looks_like_question()
    currently accepts any string > 5 chars, including decimals like "25.399..."
    
    Expected: Should return 0 questions from 'Financials' sheet
    Actual: Parser incorrectly returns numeric cells as questions
    """
    wb = Workbook()
    
    # Sheet 1: Real questions
    ws1 = wb.active
    ws1.title = "Questions"
    ws1.append(["#", "Question", "Answer"])
    ws1.append([1, "What is your company name?", ""])
    ws1.append([2, "How many employees do you have?", ""])
    
    # Sheet 2: Pure numeric data (SHOULD be skipped entirely)
    ws2 = wb.create_sheet("Financials")
    ws2.append(["Year", "Revenue", "EBITDA", "Margin %"])
    ws2.append([2020, 150000000, 22500000, 15.0])
    ws2.append([2021, 175000000, 28000000, 16.0])
    ws2.append([2022, 195000000, 33150000, 17.0])
    ws2.append([2023, 210000000, 37800000, 18.0])
    ws2.append([2024, 235000000, 44650000, 19.0])
    
    # Sheet 3: Mixed numeric with a few text cols (also should be skipped)
    ws3 = wb.create_sheet("Scoring Matrix")
    ws3.append(["Criterion", "Weight", "Score", "Weighted Score"])
    ws3.append(["Price", 30, 85, 25.5])
    ws3.append(["Quality", 25, 90, 22.5])
    ws3.append(["Experience", 20, 75, 15.0])
    ws3.append(["Innovation", 15, 80, 12.0])
    ws3.append(["References", 10, 95, 9.5])
    
    path = os.path.join(OUT_DIR, "test1_numeric_sheets_bug.xlsx")
    wb.save(path)
    print(f"Created: {path}")
    return path


def test2_corrupted_file():
    """
    TEST 2 (Easy - error handling): File with .xlsx extension but
    contains garbage bytes. Tests that:
    - API returns proper 422 error (not 500)
    - Frontend displays the error message clearly
    - Temp file is cleaned up
    """
    path = os.path.join(OUT_DIR, "test2_corrupted_not_excel.xlsx")
    # Write random binary garbage that is NOT a valid ZIP/Excel
    with open(path, "wb") as f:
        f.write(b"THIS IS NOT AN EXCEL FILE\x00\x01\x02\x03" * 100)
    print(f"Created: {path}")
    
    # Also create one that's a partial ZIP (starts with PK but truncated)
    path2 = os.path.join(OUT_DIR, "test2b_truncated_excel.xlsx")
    with open(path2, "wb") as f:
        f.write(b"PK\x03\x04" + b"\x00" * 50)  # ZIP header but truncated
    print(f"Created: {path2}")
    
    return path


def test3_writer_crash_mismatched_sheets():
    """
    TEST 3 (Medium): File where the parser extracts questions, then the 
    writer tries to write answers back but encounters issues:
    - Sheet names with characters that could break file operations
    - Very large existing content in answer cells
    - Cells with formulas
    
    The writer may crash when max_column is very large or when trying to 
    access cells beyond the sheet's dimensions.
    """
    wb = Workbook()
    
    # Sheet with many columns (wide sheet) - answer col detection may fail
    ws1 = wb.active
    ws1.title = "Wide Format RFI"
    # 20 columns of headers
    headers = ["Ref", "Category", "Sub-category", "Priority", "Question", 
               "Instructions", "Max Length", "Format", "Mandatory?", "Dependencies",
               "Answer", "Evidence Required", "Attachments", "Reviewer", "Status",
               "Score", "Comments", "Last Updated", "Version", "Notes"]
    ws1.append(headers)
    ws1.append(["Q1", "General", "Overview", "High", 
                "Provide a comprehensive overview of your organization including history, mission, and strategic direction",
                "Max 500 words", 500, "Free text", "Yes", "None",
                "", "Yes", "", "John Smith", "Pending",
                "", "", "2025-01-15", "1.0", ""])
    ws1.append(["Q2", "General", "Structure", "High",
                "Describe your organizational structure and governance model",
                "Include org chart", 300, "Free text", "Yes", "Q1",
                "", "Yes", "", "Jane Doe", "Pending",
                "", "", "2025-01-15", "1.0", ""])
    ws1.append(["Q3", "Technical", "Security", "Critical",
                "What security certifications do you hold?",
                "List all", 200, "List", "Yes", "None",
                "", "Yes - copies of certs", "", "IT Lead", "Pending",
                "", "", "2025-01-15", "1.0", ""])
    
    # Sheet with very long pre-existing answers (tests writer overwrite behavior)
    ws2 = wb.create_sheet("Pre-filled Answers")
    ws2.append(["Question", "Answer"])
    ws2.append(["What is your company name?", "A" * 10000])  # 10K char existing answer
    ws2.append(["Describe your services", "B" * 10000])
    ws2.append(["What is your headcount?", "C" * 10000])
    
    # Sheet with special characters in name (no / allowed by Excel)
    ws3 = wb.create_sheet("Section 3-4 — Overview (Draft)")
    ws3.append(["#", "Question", "Response"])
    ws3.append([1, "What is your approach to project management?", ""])
    ws3.append([2, "Describe your quality assurance process", ""])
    
    path = os.path.join(OUT_DIR, "test3_writer_stress.xlsx")
    wb.save(path)
    print(f"Created: {path}")
    return path


def test4_xlsm_macro_file():
    """
    TEST 4 (Hard): .xlsm file — the parser opens with read_only=True
    and the writer uses keep_vba=True. But we can't easily create a real
    .xlsm with openpyxl (it doesn't write VBA). Instead, create a valid
    .xlsx and rename to .xlsm to test the code path.
    
    The writer does:
      is_macro_file = ext.lower() == ".xlsm"
      wb = openpyxl.load_workbook(output_path, keep_vba=is_macro_file)
    
    But if the file was originally .xlsx renamed to .xlsm, openpyxl may
    handle keep_vba differently. Also, read_only=True mode has limitations.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Vendor Questions"
    ws.append(["No.", "RFI Question", "Vendor Response", "Max Score"])
    ws.append([1, "Describe your experience in oncology market access", "", 20])
    ws.append([2, "What is your approach to real-world evidence generation?", "", 15])
    ws.append([3, "List your key pharma client relationships (non-confidential)", "", 10])
    ws.append([4, "Describe your data analytics capabilities", "", 15])
    ws.append([5, "What AI/ML tools do you use in your workflow?", "", 10])
    
    # Save as .xlsm extension (openpyxl will still write xlsx format internally)
    path = os.path.join(OUT_DIR, "test4_fake_macro.xlsm")
    wb.save(path)
    print(f"Created: {path}")
    return path


def test5_zero_questions_and_empty_session():
    """
    TEST 5 (Complex): Files that result in edge cases for the full pipeline:
    - File with content but NO parseable questions → 422 error
    - What happens if fill-mock is called on a session with 0 questions?
    - What happens with download on an unfilled session?
    
    Also tests: extremely long filename, unicode in filename
    """
    # File A: Only section headers, no actual questions
    wb = Workbook()
    ws = wb.active
    ws.title = "Contents"
    ws.append(["Section", "Page"])
    ws.append(["1. Introduction", "3"])
    ws.append(["2. Scope", "5"])
    ws.append(["3. Requirements", "8"])
    ws.append(["4. Timeline", "12"])
    ws.append(["5. Appendices", "15"])
    
    ws2 = wb.create_sheet("Cover")
    ws2.append(["RFI DOCUMENT"])
    ws2.append(["Confidential"])
    
    path_a = os.path.join(OUT_DIR, "test5a_no_questions.xlsx")
    wb.save(path_a)
    print(f"Created: {path_a}")
    
    # File B: Very long filename with unicode
    wb2 = Workbook()
    ws = wb2.active
    ws.title = "Questions"
    ws.append(["#", "Question", "Answer"])
    ws.append([1, "What is your company name?", ""])
    
    long_name = "RFI_Questionnaire_Très_Spécial_für_Ünternehmen_日本語テスト_" + "x" * 100 + ".xlsx"
    path_b = os.path.join(OUT_DIR, long_name)
    try:
        wb2.save(path_b)
        print(f"Created: {path_b}")
    except Exception as e:
        print(f"Could not create long filename: {e}")
        path_b = os.path.join(OUT_DIR, "test5b_unicode_name_très_spécial.xlsx")
        wb2.save(path_b)
        print(f"Created: {path_b}")
    
    return path_a


if __name__ == "__main__":
    print("Generating aggressive test files...")
    test1_numeric_only_sheets_bug()
    test2_corrupted_file()
    test3_writer_crash_mismatched_sheets()
    test4_xlsm_macro_file()
    test5_zero_questions_and_empty_session()
    print("\nAll aggressive test files generated!")
