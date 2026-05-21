"""
Generate 5 test Excel files to stress-test the RFI parser + API pipeline.
Severity: Easy → Complex edge cases.
"""

import os
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

OUT_DIR = os.path.dirname(os.path.abspath(__file__))


def test1_multi_sheet_rfi():
    """
    TEST 1 (Easy): Multi-sheet Excel with different RFI sections.
    Common real-world scenario — many RFIs have 4-6 sheets.
    Potential issue: parser may fail to detect columns across sheets with different layouts.
    """
    wb = Workbook()

    # Sheet 1: Company Overview (standard layout)
    ws1 = wb.active
    ws1.title = "Company Overview"
    ws1.append(["#", "Question", "Answer"])
    ws1.append([1, "What is your company's full legal name?", ""])
    ws1.append([2, "What year was your company founded?", ""])
    ws1.append([3, "How many employees does your company have globally?", ""])

    # Sheet 2: Compliance (different column layout - no # column)
    ws2 = wb.create_sheet("Compliance")
    ws2.append(["Requirement", "Response", "Evidence"])
    ws2.append(["Do you have a written anti-bribery policy?", "", ""])
    ws2.append(["Are employees trained on data protection annually?", "", ""])
    ws2.append(["Describe your whistleblower reporting mechanism", "", ""])

    # Sheet 3: Data Security (question col is C, not B)
    ws3 = wb.create_sheet("Data Security")
    ws3.append(["Category", "Ref", "Question", "Vendor Response"])
    ws3.append(["Encryption", "DS-1", "Do you encrypt data at rest?", ""])
    ws3.append(["Encryption", "DS-2", "What encryption algorithm do you use?", ""])
    ws3.append(["Access", "DS-3", "Describe your access control methodology", ""])

    # Sheet 4: Very small sheet (should be skipped or handled gracefully)
    ws4 = wb.create_sheet("Instructions")
    ws4.append(["Please fill in all tabs"])
    ws4.append(["Return by 30 June 2025"])

    path = os.path.join(OUT_DIR, "test1_multi_sheet.xlsx")
    wb.save(path)
    print(f"Created: {path}")
    return path


def test2_empty_rows_and_merged_headers():
    """
    TEST 2 (Easy-Medium): Excel with empty rows scattered throughout,
    merged cells in the header, and inconsistent formatting.
    Common issue: parsers often crash on merged cells or miscount rows.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Vendor Assessment"

    # Merged header area (simulated — just wide text)
    ws["A1"] = "VENDOR ASSESSMENT QUESTIONNAIRE"
    ws["A2"] = "Client: Pfizer | Date: 2025-03-15"
    ws["A3"] = ""  # blank row

    # Actual header row
    ws["A4"] = "No."
    ws["B4"] = "Question"
    ws["C4"] = "Answer"
    ws["D4"] = "Comments"

    # Data with scattered empty rows
    ws["A5"] = 1
    ws["B5"] = "What is your company name?"
    ws["A6"] = 2
    ws["B6"] = "Describe your quality management system"
    ws["A7"] = ""  # empty row
    ws["B7"] = ""
    ws["A8"] = 3
    ws["B8"] = "What certifications do you hold?"
    ws["A9"] = ""  # another empty row
    ws["B9"] = ""
    ws["A10"] = 4
    ws["B10"] = "Do you have ISO 27001 certification?"
    ws["A11"] = ""  # empty
    ws["A12"] = ""  # empty
    ws["A13"] = 5
    ws["B13"] = "Please provide your business continuity plan overview"

    # Merge some cells in header
    ws.merge_cells("A1:D1")
    ws.merge_cells("A2:D2")

    path = os.path.join(OUT_DIR, "test2_empty_rows_merged.xlsx")
    wb.save(path)
    print(f"Created: {path}")
    return path


def test3_no_clear_headers():
    """
    TEST 3 (Medium): Excel with NO clear header row — data starts at row 1.
    Parser must use heuristic/fallback detection.
    Also includes very long question text and special characters.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "ESG Questions"

    # No header row at all — data starts immediately
    questions = [
        ("1.1", "Please describe your organization's environmental sustainability policy including any carbon neutrality targets, renewable energy usage, waste management protocols, and circular economy initiatives. Include quantitative metrics where possible.", ""),
        ("1.2", "What is your Scope 1, 2, and 3 greenhouse gas emissions for FY2024 (in tCO₂e)?", ""),
        ("1.3", "Do you publish an annual CSR/ESG report? If yes, provide the URL.", ""),
        ("2.1", "Describe your Diversity, Equity & Inclusion (DE&I) strategy — include gender pay gap data if available.", ""),
        ("2.2", "What % of your board members are from underrepresented groups?", ""),
        ("3.1", 'Does your company have a "modern slavery" statement per the UK Modern Slavery Act 2015?', ""),
        ("3.2", "List any ESG-related controversies or fines in the past 3 years (€/£/$)", ""),
        ("3.3", "Provide details on your supply chain due diligence process — how do you ensure Tier 2+ suppliers meet ESG standards?", ""),
    ]

    for q in questions:
        ws.append(list(q))

    path = os.path.join(OUT_DIR, "test3_no_headers_special_chars.xlsx")
    wb.save(path)
    print(f"Created: {path}")
    return path


def test4_password_protected_and_hidden_sheets():
    """
    TEST 4 (Hard): Excel with hidden sheets and sheet-level protection.
    Simulates enterprise RFIs that often ship with protected/hidden instruction sheets.
    openpyxl can read hidden sheets but protection may cause write failures.
    """
    wb = Workbook()

    # Visible data sheet
    ws1 = wb.active
    ws1.title = "RFI Questions"
    ws1.append(["#", "Question", "Your Response"])
    ws1.append([1, "What is your annual revenue?", ""])
    ws1.append([2, "How many offices do you operate globally?", ""])
    ws1.append([3, "Describe your key therapeutic area expertise", ""])
    ws1.append([4, "Do you have experience in oncology medical communications?", ""])

    # Hidden sheet (internal scoring)
    ws2 = wb.create_sheet("Scoring (Internal)")
    ws2.append(["Question ID", "Weight", "Max Score"])
    ws2.append([1, 10, 100])
    ws2.append([2, 5, 50])
    ws2.append([3, 20, 200])
    ws2.append([4, 15, 150])
    ws2.sheet_state = "hidden"

    # Another hidden sheet
    ws3 = wb.create_sheet("Admin Notes")
    ws3.append(["Do not share with vendor"])
    ws3.append(["Deadline: 15 July 2025"])
    ws3.sheet_state = "hidden"

    # Protected sheet (questions locked, answer cells unlocked)
    ws4 = wb.create_sheet("Protected Section")
    ws4.append(["Topic", "Question", "Response"])
    ws4.append(["Financial", "What is your EBITDA margin?", ""])
    ws4.append(["Financial", "Provide 3 years of audited financials", ""])
    ws4.append(["Operational", "What is your staff utilization rate?", ""])
    ws4.protection.sheet = True
    ws4.protection.password = "rfidemo"

    path = os.path.join(OUT_DIR, "test4_hidden_protected_sheets.xlsx")
    wb.save(path)
    print(f"Created: {path}")
    return path


def test5_extremely_large_and_malformed():
    """
    TEST 5 (Complex/Rare): Excel with:
    - 500+ rows (large RFI)
    - Mixed content: some cells have formulas (as strings), hyperlinks, line breaks
    - Sheet with ONLY numeric data (should be skipped)
    - Sheet with questions in rows (transposed layout — horizontal)
    - Unicode/emoji in cells
    """
    wb = Workbook()

    # Sheet 1: Large standard RFI (200 questions)
    ws1 = wb.active
    ws1.title = "Full Assessment"
    ws1.append(["Section", "#", "Question", "Response", "Notes"])
    sections = ["Governance", "Operations", "Technology", "People", "Finance"]
    for i in range(1, 201):
        section = sections[i % len(sections)]
        ws1.append([
            section,
            f"{(i // 40) + 1}.{i % 40 + 1}",
            f"Please describe your approach to {section.lower()} requirement #{i}. "
            f"Include relevant policies, procedures, and metrics.",
            "",
            ""
        ])

    # Sheet 2: Numeric data only (should be skipped — no real questions)
    ws2 = wb.create_sheet("Financial Data")
    ws2.append(["Year", "Revenue ($M)", "EBITDA ($M)", "Headcount"])
    for year in range(2018, 2026):
        ws2.append([year, year * 1.5 - 2900, year * 0.3 - 580, year - 1800])

    # Sheet 3: Transposed layout (questions as columns, not rows)
    ws3 = wb.create_sheet("Quick Facts")
    # Row 1 = questions as column headers
    ws3.append([
        "Company Name",
        "Year Founded",
        "Headquarters",
        "Number of Employees",
        "Annual Revenue",
        "Key Therapeutic Areas",
    ])
    # Row 2 = blank answers
    ws3.append(["", "", "", "", "", ""])

    # Sheet 4: Unicode, emojis, line breaks
    ws4 = wb.create_sheet("Diversity & Inclusion 🌍")
    ws4.append(["Q#", "Question", "Answer"])
    ws4.append(["D1", "What is your gender pay gap ratio? 📊", ""])
    ws4.append(["D2", "Describe your D&I initiatives:\n- Mentoring\n- ERGs\n- Training", ""])
    ws4.append(["D3", "Do you track ethnicity data? (Y/N) ✅❌", ""])
    ws4.append(["D4", "Provide your most recent WGEA report (Australia) or equivalent — résumé acceptable", ""])
    ws4.append(["D5", "名前を入力してください (Japanese: Enter your name)", ""])
    ws4.append(["D6", "¿Cuántos empleados tiene en América Latina?", ""])

    # Sheet 5: Cells with formula-like strings
    ws5 = wb.create_sheet("Formulas & Edge Cases")
    ws5.append(["#", "Question", "Response"])
    ws5.append([1, "=SUM(A1:A10) — is this parsed correctly?", ""])
    ws5.append([2, "What is your NPS score (on a scale of -100 to +100)?", ""])
    ws5.append([3, "Provide the URL: https://www.example.com/rfi?param=value&foo=bar", ""])
    ws5.append([4, "", ""])  # blank question
    ws5.append([5, "     ", ""])  # whitespace-only question
    ws5.append([6, "A", ""])  # too-short question
    ws5.append([7, "Describe your approach to AI/ML including any use of LLMs (GPT-4, Claude, etc.)", ""])

    path = os.path.join(OUT_DIR, "test5_large_malformed_unicode.xlsx")
    wb.save(path)
    print(f"Created: {path}")
    return path


if __name__ == "__main__":
    print("Generating test files...")
    test1_multi_sheet_rfi()
    test2_empty_rows_and_merged_headers()
    test3_no_clear_headers()
    test4_password_protected_and_hidden_sheets()
    test5_extremely_large_and_malformed()
    print("\nAll test files generated!")
