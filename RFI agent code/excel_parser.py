"""
Excel RFI parser — reads .xlsx/.xlsm files and extracts Q&A pairs.
Handles the wide structural variation seen across RFIs from different clients.
"""

from __future__ import annotations

import os
import re
from dataclasses import dataclass, field, asdict
from typing import Optional
import openpyxl
from openpyxl.utils import get_column_letter


@dataclass
class RFIQuestion:
    sheet_name: str
    row: int
    question_col: str  # column letter
    answer_col: Optional[str]  # column letter, if detected
    question_number: str  # e.g. "1.1", "2", "A.3"
    question_text: str
    existing_answer: str  # what's already filled in
    category_hint: str  # from sheet name, if available


# Common header patterns for question and answer columns
QUESTION_HEADERS = re.compile(
    r"^(questions?\b|q\b|description|requirement|criteria|query|item|topic|details|please\s)",
    re.IGNORECASE,
)
QUESTION_LABEL_ONLY = re.compile(
    r"^(questions?\s*#?|q\s*#?)$",
    re.IGNORECASE,
)
ANSWER_HEADERS = re.compile(
    r"^(answer|response|reply|draft|your answer|vendor|supplier|agency|avalere|comments?$)",
    re.IGNORECASE,
)
ANSWER_EXCLUDE = re.compile(
    r"(owner|assigned|responsible|lead|status)",
    re.IGNORECASE,
)
NUMBER_HEADERS = re.compile(
    r"^(#|no\.?|ref\.?|id|number|question\s*#|q#|s\.?no)",
    re.IGNORECASE,
)
SKIP_HEADERS = re.compile(
    r"^(assigned|owner|response owner|status|notes|weight|score|max score|internal)",
    re.IGNORECASE,
)


def _cell_text(cell_value) -> str:
    """Convert cell value to clean string."""
    if cell_value is None:
        return ""
    return str(cell_value).strip()


def _is_section_header(text: str) -> bool:
    """Check if a row is a section header rather than a question."""
    if not text:
        return False
    # Section headers are typically short, don't end with '?', and often
    # match patterns like "Section 1:", "A. Conflicts", "1. Company Details"
    if len(text) < 80 and not text.endswith("?"):
        if re.match(r"^(section\s+\d|part\s+\d|\d+\.\s+[A-Z]|[A-Z]\.\s+[A-Z])", text, re.IGNORECASE):
            return True
    return False


def _looks_like_question(text: str) -> bool:
    """Heuristic: does this text look like a question, prompt, or field label?"""
    if not text or len(text) < 5:
        return False
    # Reject pure numeric values (integers, floats, percentages, currency)
    stripped = text.strip().rstrip("%")
    try:
        float(stripped.replace(",", "").replace("$", "").replace("£", "").replace("€", ""))
        return False
    except ValueError:
        pass
    # Ends with ? — definitely a question
    if text.rstrip().endswith("?"):
        return True
    # Request/instruction starters
    question_starters = (
        "please", "provide", "describe", "what", "how", "do you", "does your",
        "have you", "are you", "is your", "can you", "will you", "list",
        "explain", "confirm", "indicate", "specify", "outline", "detail",
        "share", "give", "state", "name", "select", "choose"
    )
    lower = text.lower().lstrip("0123456789.-) ")
    if any(lower.startswith(s) for s in question_starters):
        return True
    # Short field labels are also valid (e.g., "Company Website", "Office Address")
    # Accept anything > 5 chars that isn't just a number or section header
    # Require at least 2 words for the catch-all (single words are category labels, not questions)
    if len(text) > 5 and not _is_section_header(text) and " " in text.strip():
        return True
    return False


def _extract_question_number(row_data: list, number_col: int | None) -> str:
    """Extract question number from row."""
    if number_col is not None:
        val = _cell_text(row_data[number_col])
        if val:
            return val
    return ""


def _detect_columns(sheet) -> tuple:
    """
    Detect which columns contain: number, question, answer, and the header row index.
    Returns (number_col, question_col, answer_col, header_row_idx).
    """
    # Scan first 10 rows for headers
    all_rows = []
    for row_idx, row in enumerate(sheet.iter_rows(max_row=10, values_only=True)):
        texts = [_cell_text(c) for c in row]
        all_rows.append((row_idx, texts))

    # Find the best header row — the row with the most recognized column headers
    best_header = None
    best_score = 0

    for row_idx, texts in all_rows:
        q_cols = [i for i, t in enumerate(texts) if t and QUESTION_HEADERS.search(t)]
        a_cols = [i for i, t in enumerate(texts) if t and ANSWER_HEADERS.search(t) and not ANSWER_EXCLUDE.search(t)]
        n_cols = [i for i, t in enumerate(texts) if t and NUMBER_HEADERS.search(t)]
        score = len(q_cols) + len(a_cols) + len(n_cols)

        if score > best_score:
            best_score = score
            best_header = (row_idx, n_cols, q_cols, a_cols, texts)

    if best_header and best_score >= 2:
        row_idx, n_cols, q_cols, a_cols, header_texts = best_header
        number_col = n_cols[0] if n_cols else None

        # Trust the header labels — they tell us exactly which column is which
        # If question_col collides with number_col, pick the next q_col candidate
        question_col = q_cols[0] if q_cols else None
        if question_col is not None and question_col == number_col and len(q_cols) > 1:
            question_col = q_cols[1]
        answer_col = a_cols[0] if a_cols else None

        # If the detected question column header is just a label ("Questions", "Question #")
        # rather than actual question content, the real questions are likely in the next column.
        # BUT: only apply this heuristic if we don't already have a confirmed answer column,
        # because the next column IS the answer column when both headers are detected.
        if question_col is not None and answer_col is None:
            header_text = header_texts[question_col] if question_col < len(header_texts) else ""
            if QUESTION_LABEL_ONLY.match(header_text.strip()):
                # Check if the next column has longer text in data rows (= actual questions)
                next_col = question_col + 1
                if next_col < len(header_texts):
                    data_lens_q = []
                    data_lens_next = []
                    for dr_idx, dr_texts in all_rows:
                        if dr_idx <= row_idx:
                            continue
                        data_lens_q.append(len(dr_texts[question_col]) if question_col < len(dr_texts) else 0)
                        data_lens_next.append(len(dr_texts[next_col]) if next_col < len(dr_texts) else 0)
                    avg_q = sum(data_lens_q) / max(len(data_lens_q), 1)
                    avg_next = sum(data_lens_next) / max(len(data_lens_next), 1)
                    if avg_next > avg_q:
                        # Also check if question_col looks like a number column
                        if number_col is None:
                            number_col = question_col
                        question_col = next_col

        # If no question header found but we have an answer header, question is likely
        # the column before the answer (or the longest-text column that isn't number/answer/skip)
        if question_col is None and answer_col is not None:
            skip_cols = set()
            if number_col is not None:
                skip_cols.add(number_col)
            skip_cols.add(answer_col)
            for i, t in enumerate(header_texts):
                if t and SKIP_HEADERS.search(t):
                    skip_cols.add(i)
            # Pick the non-skipped column with longest average text in data rows
            col_avg = {}
            for dr_idx, dr_texts in all_rows:
                if dr_idx <= row_idx:
                    continue
                for ci, t in enumerate(dr_texts):
                    if ci in skip_cols:
                        continue
                    if ci not in col_avg:
                        col_avg[ci] = []
                    col_avg[ci].append(len(t))
            if col_avg:
                avg_lens = {ci: sum(lens) / len(lens) for ci, lens in col_avg.items()}
                question_col = max(avg_lens, key=avg_lens.get)
            else:
                question_col = max(0, (answer_col or 1) - 1)

        # If we still don't have a question column, default to col B (index 1)
        if question_col is None:
            question_col = 1

        return number_col, question_col, answer_col, row_idx

    # Fallback: no clear headers. Find the column with the longest average text.
    col_text_lens = {}
    for row in sheet.iter_rows(min_row=1, max_row=15, values_only=True):
        for i, cell in enumerate(row):
            t = _cell_text(cell)
            if i not in col_text_lens:
                col_text_lens[i] = []
            col_text_lens[i].append(len(t))

    if col_text_lens:
        avg_lens = {ci: sum(lens) / len(lens) for ci, lens in col_text_lens.items()}
        question_col = max(avg_lens, key=avg_lens.get)
        # Answer is the next-longest column
        candidates = sorted(
            [(ci, al) for ci, al in avg_lens.items() if ci != question_col and al > 5],
            key=lambda x: (-x[1], x[0])
        )
        answer_col = candidates[0][0] if candidates else (question_col + 1)
        # Number col is typically the first column if it has short entries
        number_col = 0 if avg_lens.get(0, 0) < 10 and 0 != question_col else None
        return number_col, question_col, answer_col, 0

    return None, 1, 2, 0


def _category_from_sheet_name(sheet_name: str) -> str:
    """Infer category hint from sheet name."""
    name = sheet_name.lower().strip()

    mappings = {
        "company": "Company Information",
        "overview": "Company Information",
        "general info": "Company Information",
        "compliance": "Compliance",
        "code of conduct": "Compliance",
        "amendments": "Compliance",
        "legal": "Legal",
        "provision": "Legal",
        "data": "Data & Information Security",
        "security": "Data & Information Security",
        "information security": "Data & Information Security",
        "privacy": "Data & Information Security",
        "gdpr": "Data & Information Security",
        "sustainab": "ESG",
        "esg": "ESG",
        "environment": "ESG",
        "social": "ESG",
        "governance": "ESG",
        "people": "People Information",
        "staff": "People Information",
        "team": "People Information",
        "resource": "People Information",
        "hr": "People Information",
        "supplier": "Suppliers & Freelancers",
        "freelance": "Suppliers & Freelancers",
        "subcontract": "Suppliers & Freelancers",
        "vendor": "Suppliers & Freelancers",
        "technolog": "Technology & AI",
        "ai": "Technology & AI",
        "digital": "Technology & AI",
        "commercial": "Commercial Information",
        "financial": "Commercial Information",
        "service": "Commercial Information",
        "capabilities": "Commercial Information",
        "pricing": "Commercial Information",
    }

    for keyword, category in mappings.items():
        if keyword in name:
            return category

    return ""  # No hint from sheet name


def parse_rfi(filepath: str) -> list[RFIQuestion]:
    """
    Parse an RFI Excel file and extract all Q&A pairs.
    Returns a list of RFIQuestion objects.
    """
    ext = os.path.splitext(filepath)[1].lower()
    read_only = ext == ".xlsm"  # read-only for macro files to preserve them

    wb = openpyxl.load_workbook(filepath, read_only=read_only, data_only=True)
    questions = []

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]

        # Skip very small sheets (likely instruction or metadata)
        if ws.max_row is None or ws.max_row < 3:
            continue

        number_col, question_col, answer_col, header_row = _detect_columns(ws)
        category_hint = _category_from_sheet_name(sheet_name)

        # Read all rows after the header
        current_section = ""
        for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
            if row_idx <= header_row:
                continue

            cells = list(row)
            q_text = _cell_text(cells[question_col]) if question_col < len(cells) else ""
            a_text = _cell_text(cells[answer_col]) if answer_col is not None and answer_col < len(cells) else ""
            q_num = _extract_question_number(cells, number_col) if number_col is not None and number_col < len(cells) else ""

            # Skip empty rows
            if not q_text:
                continue

            # Track section headers but don't add them as questions
            if _is_section_header(q_text):
                current_section = q_text
                continue

            # Skip if this doesn't look like a real question
            if not _looks_like_question(q_text):
                continue

            questions.append(RFIQuestion(
                sheet_name=sheet_name,
                row=row_idx + 1,  # 1-indexed for Excel
                question_col=get_column_letter(question_col + 1),
                answer_col=get_column_letter(answer_col + 1) if answer_col is not None else None,
                question_number=q_num,
                question_text=q_text,
                existing_answer=a_text,
                category_hint=category_hint or current_section,
            ))

    wb.close()
    return questions


def extract_client_from_filename(filename: str) -> str:
    """Try to extract client name from RFI filename."""
    known_clients = [
        "Pfizer", "Gilead", "AbbVie", "AstraZeneca", "AZ", "Novartis",
        "Servier", "UCB", "GSK", "Chiesi", "Rigel", "Exact Sciences",
    ]
    basename = os.path.basename(filename)
    for client in known_clients:
        if client.lower() in basename.lower():
            # Normalize AZ -> AstraZeneca
            if client == "AZ":
                return "AstraZeneca"
            return client
    return ""


if __name__ == "__main__":
    import json

    script_dir = os.path.dirname(os.path.abspath(__file__))
    rfi_dir = os.path.join(script_dir, "..")
    files = [f for f in os.listdir(rfi_dir) if f.endswith((".xlsx", ".xlsm"))]

    print(f"Found {len(files)} RFI files\n")

    for fname in files[:3]:
        path = os.path.join(rfi_dir, fname)
        client = extract_client_from_filename(fname)
        questions = parse_rfi(path)
        print(f"=== {fname[:60]} ===")
        print(f"  Client: {client or '(unknown)'}")
        print(f"  Questions found: {len(questions)}")
        for q in questions[:3]:
            print(f"  [{q.sheet_name}] R{q.row} {q.question_number}: {q.question_text[:70]}...")
            if q.existing_answer:
                print(f"    Answer: {q.existing_answer[:70]}...")
            print(f"    Category hint: {q.category_hint or '(none)'}")
        print()
