"""
Excel writer — creates a filled copy of the RFI with answers, confidence scores, and color-coding.
"""

from __future__ import annotations

import os
import shutil
from typing import Optional

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter


# Color scheme for confidence levels
FILL_GREEN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")   # ≥ 0.80
FILL_YELLOW = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # 0.50–0.79
FILL_RED = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")     # < 0.50
FILL_HEADER = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
FONT_HEADER = Font(color="FFFFFF", bold=True)


def _confidence_fill(confidence: float) -> PatternFill:
    if confidence >= 0.80:
        return FILL_GREEN
    elif confidence >= 0.50:
        return FILL_YELLOW
    else:
        return FILL_RED


def write_filled_rfi(
    source_path: str,
    questions: list[dict],
    output_dir: Optional[str] = None,
) -> str:
    """
    Create a filled copy of the RFI Excel file.

    - Writes generated answers into the answer column
    - Adds a "Confidence" column with scores
    - Color-codes answer cells by confidence level
    - Returns the path to the output file
    """
    basename = os.path.splitext(os.path.basename(source_path))[0]
    ext = os.path.splitext(source_path)[1]
    out_dir = output_dir or os.path.dirname(source_path)
    output_path = os.path.join(out_dir, f"{basename}_FILLED{ext}")

    # Copy the original file to preserve all formatting, macros, etc.
    shutil.copy2(source_path, output_path)

    # Open the copy for editing (NOT read-only)
    # keep_vba=True only for .xlsm files (preserves macros);
    # .xlsx files must NOT contain VBA or Excel will reject them
    is_macro_file = ext.lower() == ".xlsm"
    wb = openpyxl.load_workbook(output_path, keep_vba=is_macro_file)

    # Group questions by sheet
    by_sheet = {}
    for q in questions:
        sheet = q["sheet_name"]
        if sheet not in by_sheet:
            by_sheet[sheet] = []
        by_sheet[sheet].append(q)

    for sheet_name, sheet_qs in by_sheet.items():
        if sheet_name not in wb.sheetnames:
            continue

        ws = wb[sheet_name]

        # Standardized output layout (always left to right):
        # [existing columns] → AI Answer → AI Confidence → Citation
        # Place new columns after the last used column to avoid overwriting
        last_col = ws.max_column or 1
        ans_col_num = last_col + 1
        conf_col_num = last_col + 2
        cite_col_num = last_col + 3

        ans_col_letter = get_column_letter(ans_col_num)
        conf_col_letter = get_column_letter(conf_col_num)
        cite_col_letter = get_column_letter(cite_col_num)

        # Write headers
        min_row = min(q["row"] for q in sheet_qs)
        header_row = max(1, min_row - 1)

        try:
            h = ws[f"{ans_col_letter}{header_row}"]
            h.value = "AI Answer"
            h.fill = FILL_HEADER
            h.font = FONT_HEADER
            h.alignment = Alignment(horizontal="left", wrap_text=True)
        except (AttributeError, TypeError):
            pass

        try:
            h = ws[f"{conf_col_letter}{header_row}"]
            h.value = "AI Confidence"
            h.fill = FILL_HEADER
            h.font = FONT_HEADER
            h.alignment = Alignment(horizontal="center")
        except (AttributeError, TypeError):
            pass

        try:
            h = ws[f"{cite_col_letter}{header_row}"]
            h.value = "Citation"
            h.fill = FILL_HEADER
            h.font = FONT_HEADER
            h.alignment = Alignment(horizontal="left")
        except (AttributeError, TypeError):
            pass

        # Write data rows: Answer → Confidence → Citation
        for q in sheet_qs:
            row = q["row"]
            answer = q.get("generated_answer", "")
            confidence = q.get("confidence") or 0
            citation = q.get("citation", "")

            try:
                # AI Answer
                ans_cell = ws[f"{ans_col_letter}{row}"]
                ans_cell.value = answer or ""
                ans_cell.fill = _confidence_fill(confidence)
                ans_cell.alignment = Alignment(wrap_text=True)

                # AI Confidence
                conf_cell = ws[f"{conf_col_letter}{row}"]
                conf_cell.value = round(confidence, 2)
                conf_cell.fill = _confidence_fill(confidence)
                conf_cell.alignment = Alignment(horizontal="center")
                conf_cell.number_format = "0%"

                # Citation
                cite_cell = ws[f"{cite_col_letter}{row}"]
                cite_cell.value = citation or ""
                cite_cell.alignment = Alignment(horizontal="left", wrap_text=True)
            except (AttributeError, TypeError):
                continue

    wb.save(output_path)
    wb.close()

    return output_path


def write_summary(
    questions: list[dict],
    output_path: str,
    source_filename: str,
    client_name: str = "",
) -> str:
    """
    Generate a markdown summary report of the fill results.
    """
    total = len(questions)
    filled = sum(1 for q in questions if q.get("generated_answer") and not q["generated_answer"].startswith("["))
    high_conf = sum(1 for q in questions if q.get("confidence", 0) >= 0.80)
    medium_conf = sum(1 for q in questions if 0.50 <= q.get("confidence", 0) < 0.80)
    low_conf = sum(1 for q in questions if q.get("confidence", 0) < 0.50)
    flagged = [q for q in questions if q.get("review_flag")]

    # Sheet breakdown
    sheet_counts = {}
    for q in questions:
        sheet = q.get("sheet_name", "Unknown")
        if sheet not in sheet_counts:
            sheet_counts[sheet] = {"total": 0, "filled": 0}
        sheet_counts[sheet]["total"] += 1
        if q.get("generated_answer") and not q["generated_answer"].startswith("["):
            sheet_counts[sheet]["filled"] += 1

    fill_rate = (filled / total * 100) if total > 0 else 0

    lines = [
        f"# RFI Fill Summary",
        f"",
        f"- **Source:** {source_filename}",
        f"- **Client:** {client_name or 'Unknown'}",
        f"- **Total questions:** {total}",
        f"- **Filled:** {filled} ({fill_rate:.0f}%)",
        f"- **Confidence breakdown:**",
        f"  - 🟢 High (≥80%): {high_conf}",
        f"  - 🟡 Medium (50-79%): {medium_conf}",
        f"  - 🔴 Low (<50%): {low_conf}",
        f"",
        f"## By Sheet",
        f"",
        f"| Sheet | Total | Filled | Rate |",
        f"|---|---|---|---|",
    ]

    for sheet, counts in sorted(sheet_counts.items()):
        rate = (counts["filled"] / counts["total"] * 100) if counts["total"] > 0 else 0
        lines.append(f"| {sheet} | {counts['total']} | {counts['filled']} | {rate:.0f}% |")

    if flagged:
        lines.extend([
            f"",
            f"## Flagged for Review ({len(flagged)})",
            f"",
        ])
        for q in flagged:
            lines.append(f"- **[{q.get('sheet_name')}] Row {q.get('row')}**: {q['question_text'][:100]}")
            lines.append(f"  - Flag: {q['review_flag']}")
            lines.append(f"  - Confidence: {q.get('confidence', 0):.0%}")
            lines.append(f"")

    # Needs review section (low confidence, not flagged)
    needs_review = [q for q in questions if q.get("confidence", 0) < 0.50 and not q.get("review_flag")]
    if needs_review:
        lines.extend([
            f"",
            f"## Needs Manual Answer ({len(needs_review)})",
            f"",
        ])
        for q in needs_review[:20]:  # Cap at 20
            lines.append(f"- **[{q.get('sheet_name')}] Row {q.get('row')}**: {q['question_text'][:100]}")
            lines.append(f"")

    text = "\n".join(lines)
    with open(output_path, "w", encoding="utf-8") as f:
        f.write(text)

    return output_path
